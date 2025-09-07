import openpyxl

def get_columns(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    return [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

def read_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    columns = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for col, val in zip(columns, row):
            row_dict[col] = val or ""
        data.append(row_dict)
    return columns, data

def write_excel(data, file_path):
    if not data:
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    columns = list(data[0].keys())
    ws.append(columns)
    for row in data:
        ws.append([row[col] for col in columns])
    wb.save(file_path)
