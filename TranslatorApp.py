import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
from googletrans import Translator
import os

# 初始化翻译器
translator = Translator()

# 支持语言字典（可扩展）
LANGUAGES = {
    '英文': 'en',
    '日文': 'ja',
    '德文': 'de',
    '法文': 'fr',
    '中文': 'zh-cn'
}

class TranslatorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel翻译工具")
        self.file_path = ""
        self.output_dir = ""
        self.selected_languages = []

        # 文件选择
        tk.Button(root, text="选择Excel文件", command=self.choose_file).pack(pady=5)
        self.file_label = tk.Label(root, text="未选择文件")
        self.file_label.pack()

        # 语言选择
        tk.Label(root, text="选择翻译语言（多选）").pack()
        self.lang_listbox = tk.Listbox(root, selectmode=tk.MULTIPLE)
        for lang in LANGUAGES.keys():
            self.lang_listbox.insert(tk.END, lang)
        self.lang_listbox.pack()

        # 输出文件夹选择
        tk.Button(root, text="选择导出文件夹", command=self.choose_output_dir).pack(pady=5)
        self.output_label = tk.Label(root, text="未选择文件夹")
        self.output_label.pack()

        # 开始翻译按钮
        tk.Button(root, text="开始翻译", command=self.translate_file).pack(pady=10)

    def choose_file(self):
        self.file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if self.file_path:
            self.file_label.config(text=os.path.basename(self.file_path))

    def choose_output_dir(self):
        self.output_dir = filedialog.askdirectory()
        if self.output_dir:
            self.output_label.config(text=self.output_dir)

    def translate_file(self):
        if not self.file_path or not self.output_dir:
            messagebox.showerror("错误", "请先选择文件和导出文件夹")
            return

        # 获取选中语言
        selected_indices = self.lang_listbox.curselection()
        if not selected_indices:
            messagebox.showerror("错误", "请至少选择一种语言")
            return
        self.selected_languages = [self.lang_listbox.get(i) for i in selected_indices]

        # 打开Excel文件
        wb = openpyxl.load_workbook(self.file_path)
        ws = wb.active

        # 假设标题在第2列，描述在第3列
        for lang_name in self.selected_languages:
            lang_code = LANGUAGES[lang_name]
            new_wb = openpyxl.Workbook()
            new_ws = new_wb.active
            new_ws.append(['品名', '标题', '描述'])  # 表头

            for row in ws.iter_rows(min_row=2, values_only=True):
                name = row[0]
                title = row[1] or ""
                desc = row[2] or ""
                # 翻译
                translated_title = translator.translate(title, dest=lang_code).text
                translated_desc = translator.translate(desc, dest=lang_code).text
                new_ws.append([name, translated_title, translated_desc])

            output_file = os.path.join(self.output_dir, f"{os.path.splitext(os.path.basename(self.file_path))[0]}_{lang_code}.xlsx")
            new_wb.save(output_file)

        messagebox.showinfo("完成", "翻译完成！")

if __name__ == "__main__":
    root = tk.Tk()
    app = TranslatorApp(root)
    root.mainloop()
